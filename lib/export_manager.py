"""
数据导出管理器模块
提供Excel导出和附件文件复制的通用功能
"""

import os
import shutil
from typing import List, Dict, Optional
from PyQt6.QtWidgets import QMessageBox, QFileDialog, QWidget


class ExportManager:
    """通用导出管理器，负责数据导出和附件文件管理"""
    
    def __init__(self, data_storage=None):
        """
        初始化导出管理器
        
        Args:
            data_storage: 数据存储实例，用于获取附件信息
        """
        self.data_storage = data_storage
    
    def export_records_to_excel(self, 
                               records: List[Dict], 
                               field_headers: List[str], 
                               parent_widget: Optional[QWidget] = None,
                               user_name: str = None) -> bool:
        """
        导出记录到Excel文件
        
        Args:
            records: 要导出的记录列表
            field_headers: 字段标题列表
            parent_widget: 父窗口，用于显示对话框
            user_name: 用户名，用于获取附件信息
            
        Returns:
            是否导出成功
        """
        if not records:
            if parent_widget:
                QMessageBox.warning(parent_widget, "警告", "没有记录可导出！")
            return False
        
        # 确认导出对话框
        if not self._confirm_export(records, parent_widget):
            return False
        
        # 选择保存位置
        save_path = self._get_save_path(parent_widget)
        if not save_path:
            return False
        
        try:
            # 检查并安装openpyxl依赖
            if not self._ensure_openpyxl_available(parent_widget):
                return False
            
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "就诊记录"
            
            # 添加表头（包含附件字段）
            headers_with_attachment = field_headers + ["附件"]
            for col, header in enumerate(headers_with_attachment, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 添加数据行
            for row_idx, record in enumerate(records, 2):
                # 添加基本字段
                for col_idx, field_name in enumerate(self._get_field_names(), 1):
                    value = record.get(field_name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
                
                # 添加附件字段
                attachments_str = self._get_attachments_string(record, user_name)
                ws.cell(row=row_idx, column=len(field_headers) + 1, value=attachments_str)
            
            # 应用Excel格式设置
            self._apply_excel_formatting(ws, len(records) + 1, len(headers_with_attachment))
            
            # 保存Excel文件
            wb.save(save_path)
            
            # 导出附件
            if user_name:
                self._export_attachments(records, save_path, user_name, parent_widget)
            
            if parent_widget:
                QMessageBox.information(parent_widget, "成功", 
                                      f"数据已成功导出到：\n{save_path}")
            return True
            
        except Exception as e:
            if parent_widget:
                QMessageBox.critical(parent_widget, "错误", 
                                   f"导出失败：\n{str(e)}")
            return False
    
    def _confirm_export(self, records: List[Dict], parent_widget: Optional[QWidget]) -> bool:
        """
        确认导出对话框
        
        Args:
            records: 要导出的记录列表
            parent_widget: 父窗口
            
        Returns:
            用户是否确认导出
        """
        if not parent_widget:
            return True
        
        count = len(records)
        reply = QMessageBox.question(
            parent_widget,
            '确认导出',
            f'已选中 {count} 条记录，是否确认导出？',
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.Yes
        )
        
        return reply == QMessageBox.StandardButton.Yes
    
    def _get_save_path(self, parent_widget: Optional[QWidget]) -> Optional[str]:
        """
        获取保存路径
        
        Args:
            parent_widget: 父窗口
            
        Returns:
            用户选择的保存路径，如果取消则返回None
        """
        file_dialog = QFileDialog()
        file_path, _ = file_dialog.getSaveFileName(
            parent_widget,
            '选择导出位置',
            '就诊记录.xlsx',
            'Excel文件 (*.xlsx);;所有文件 (*)'
        )
        
        return file_path if file_path else None
    
    def _ensure_openpyxl_available(self, parent_widget: Optional[QWidget]) -> bool:
        """
        检查并确保openpyxl可用
        
        Args:
            parent_widget: 父窗口
            
        Returns:
            openpyxl是否可用
        """
        try:
            import openpyxl
            return True
        except ImportError:
            if parent_widget:
                reply = QMessageBox.question(
                    parent_widget,
                    '缺少依赖',
                    'Excel导出功能需要安装openpyxl库。\n是否立即安装？',
                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                    QMessageBox.StandardButton.Yes
                )
                
                if reply == QMessageBox.StandardButton.Yes:
                    return self._install_openpyxl(parent_widget)
            return False
    
    def _install_openpyxl(self, parent_widget: Optional[QWidget]) -> bool:
        """
        安装openpyxl依赖
        
        Args:
            parent_widget: 父窗口
            
        Returns:
            是否安装成功
        """
        try:
            import subprocess
            import sys
            
            # 尝试安装openpyxl
            result = subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], 
                                  capture_output=True, text=True)
            
            if result.returncode == 0:
                if parent_widget:
                    QMessageBox.information(parent_widget, "成功", "openpyxl已成功安装！")
                return True
            else:
                if parent_widget:
                    QMessageBox.warning(parent_widget, "失败", 
                                      f"安装openpyxl失败：\n{result.stderr}")
                return False
                
        except Exception as e:
            if parent_widget:
                QMessageBox.critical(parent_widget, "错误", 
                                   f"安装openpyxl时发生错误：\n{str(e)}")
            return False
    
    def _get_field_names(self) -> List[str]:
        """
        获取字段名列表（对应数据库字段）
        
        Returns:
            字段名列表
        """
        return [
            'visit_record_id', 'date', 'hospital', 'department', 'doctor',
            'organ_system', 'reason', 'diagnosis', 'medication', 'remark'
        ]
    
    def _get_attachments_string(self, record: Dict, user_name: str) -> str:
        """
        获取记录的附件字符串
        
        Args:
            record: 记录字典
            user_name: 用户名
            
        Returns:
            附件文件名字符串，用分号分隔
        """
        if not self.data_storage or not user_name:
            return ""
        
        visit_record_id = record.get('visit_record_id')
        if not visit_record_id:
            return ""
        
        try:
            attachments = self.data_storage.get_visit_attachments(user_name, visit_record_id)
            if not attachments:
                return ""
            
            # 提取文件名（不含扩展名）
            file_names = []
            for attachment in attachments:
                file_path = attachment.get('file_path', '')
                if file_path:
                    # 获取文件名（不含扩展名）
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    file_names.append(base_name)
            
            return "; ".join(file_names)
            
        except Exception as e:
            print(f"获取附件信息失败: {e}")
            return ""
    
    def _export_attachments(self, records: List[Dict], excel_path: str, 
                          user_name: str, parent_widget: Optional[QWidget]):
        """
        导出附件文件
        
        Args:
            records: 记录列表
            excel_path: Excel文件路径
            user_name: 用户名
            parent_widget: 父窗口
        """
        if not self.data_storage:
            return
        
        # 创建附件文件夹
        excel_dir = os.path.dirname(excel_path)
        excel_name = os.path.splitext(os.path.basename(excel_path))[0]
        attachment_dir = os.path.join(excel_dir, f"{excel_name}附件")
        
        # 收集所有需要复制的附件
        attachments_to_copy = []
        
        for record in records:
            visit_record_id = record.get('visit_record_id')
            if not visit_record_id:
                continue
            
            try:
                attachments = self.data_storage.get_visit_attachments(user_name, visit_record_id)
                for attachment in attachments:
                    file_path = attachment.get('file_path', '')
                    if file_path and os.path.exists(file_path):
                        attachments_to_copy.append(file_path)
            except Exception as e:
                print(f"获取记录 {visit_record_id} 的附件失败: {e}")
        
        # 如果没有附件需要复制，跳过
        if not attachments_to_copy:
            return
        
        try:
            # 创建附件目录
            if not os.path.exists(attachment_dir):
                os.makedirs(attachment_dir)
            
            # 复制附件文件
            copied_count = 0
            failed_files = []
            
            for file_path in attachments_to_copy:
                try:
                    file_name = os.path.basename(file_path)
                    dest_path = os.path.join(attachment_dir, file_name)
                    
                    # 如果目标文件已存在，添加数字后缀
                    counter = 1
                    base_name, ext = os.path.splitext(file_name)
                    while os.path.exists(dest_path):
                        new_name = f"{base_name}_{counter}{ext}"
                        dest_path = os.path.join(attachment_dir, new_name)
                        counter += 1
                    
                    shutil.copy2(file_path, dest_path)
                    copied_count += 1
                    
                except Exception as e:
                    print(f"复制文件 {file_path} 失败: {e}")
                    failed_files.append(os.path.basename(file_path))
            
            # 显示复制结果
            if parent_widget:
                if failed_files:
                    QMessageBox.warning(
                        parent_widget, 
                        "附件复制完成",
                        f"成功复制 {copied_count} 个附件到：\n{attachment_dir}\n\n"
                        f"以下文件复制失败：\n" + "\n".join(failed_files)
                    )
                else:
                    QMessageBox.information(
                        parent_widget,
                        "附件复制完成", 
                        f"成功复制 {copied_count} 个附件到：\n{attachment_dir}"
                    )
        
        except Exception as e:
            if parent_widget:
                QMessageBox.warning(parent_widget, "警告", 
                                  f"附件复制过程中发生错误：\n{str(e)}")
    
    def _apply_excel_formatting(self, worksheet, total_rows: int, total_cols: int):
        """
        应用Excel格式设置
        
        Args:
            worksheet: openpyxl工作表对象
            total_rows: 总行数（包括表头）
            total_cols: 总列数
        """
        try:
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 定义样式
            header_font = Font(bold=True)  # 表头加粗字体
            header_alignment = Alignment(horizontal='center', vertical='center')  # 表头居中对齐
            center_alignment = Alignment(horizontal='center', vertical='center')  # 左右居中+垂直居中
            wrap_center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 自动换行+垂直居中+左右居中
            wrap_alignment = Alignment(wrap_text=True, vertical='center')  # 自动换行+垂直居中
            vertical_center_alignment = Alignment(vertical='center')  # 仅垂直居中
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )  # 细框线
            
            # 1. C-K列开启自动换行（对应列索引3-11）
            wrap_columns = list(range(3, 12))  # C列到K列
            for col in wrap_columns:
                if col <= total_cols:
                    for row in range(1, total_rows + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row == 1:
                            # 表头行：加粗+居中+自动换行
                            cell.font = header_font
                            cell.alignment = wrap_center_alignment
                        else:
                            # 数据行：自动换行+垂直居中
                            if col in range(3, 7):  # C-F列需要特殊处理（步骤5中设置左右居中）
                                # C-F列的数据行处理将在步骤5中进行（自动换行+左右居中+垂直居中）
                                pass
                            else:
                                cell.alignment = wrap_alignment
            
            # 2. B-K列自动调整列宽（对应列索引2-11）
            for col in range(2, 12):  # B列到K列
                if col <= total_cols:
                    col_letter = get_column_letter(col)
                    # 计算列的最大内容宽度
                    max_length = 0
                    for row in range(1, total_rows + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value:
                            # 考虑换行符，计算实际显示宽度
                            cell_text = str(cell.value)
                            lines = cell_text.split('\n')
                            max_line_length = max(len(lines[0]) if lines else 0, 8)  # 使用第一行长度，最小8
                            max_length = max(max_length, max_line_length)
                    
                    # 设置列宽，最小宽度10，最大宽度50
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # 3. 第一行（表头）文字加粗、居中、冻结首行
            for col in range(1, total_cols + 1):
                cell = worksheet.cell(row=1, column=col)
                if col not in wrap_columns:  # 非自动换行列的表头处理
                    cell.font = header_font
                    cell.alignment = header_alignment
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'
            
            # 4. 有数据的行（包括表头）的A-K列加上所有框线
            for row in range(1, total_rows + 1):
                for col in range(1, min(12, total_cols + 1)):  # A列到K列
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
            
            # 5. A-F列文字左右居中（对应列索引1-6）
            center_columns = list(range(1, 7))  # A到F列
            for col in center_columns:
                if col <= total_cols:
                    for row in range(1, total_rows + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if row == 1:
                            # 表头行已经在步骤3中处理
                            continue
                        else:
                            # 数据行：左右居中+垂直居中
                            if col in range(3, 7):  # C-F列需要特殊处理：自动换行+左右居中+垂直居中
                                cell.alignment = wrap_center_alignment
                            else:  # A-B列：仅左右居中+垂直居中
                                cell.alignment = center_alignment
            
            # 6. 所有单元格垂直居中（处理剩余未处理的单元格）
            for row in range(1, total_rows + 1):
                for col in range(1, total_cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    # 如果单元格还没有设置对齐方式，则设置为垂直居中
                    if cell.alignment == Alignment():
                        cell.alignment = vertical_center_alignment
            
        except Exception as e:
            print(f"应用Excel格式时发生错误: {e}")


# ==================== 便捷函数 ====================

def export_health_records(records: List[Dict], 
                         field_headers: List[str],
                         parent_widget: Optional[QWidget] = None,
                         user_name: str = None,
                         data_storage=None) -> bool:
    """
    导出健康记录的便捷函数
    
    Args:
        records: 要导出的记录列表
        field_headers: 字段标题列表
        parent_widget: 父窗口
        user_name: 用户名
        data_storage: 数据存储实例
        
    Returns:
        是否导出成功
    """
    export_manager = ExportManager(data_storage)
    return export_manager.export_records_to_excel(
        records, field_headers, parent_widget, user_name
    ) 